# Excel Cell Extractor
# Function: Extract specified column data from Excel file and save to txt file
# Environment: Python 3.8+, openpyxl 3.1.2

import openpyxl
import os

def extract_excel_column(excel_path, column_name, output_path):
    """
    Extract data from a specified column in Excel file
    :param excel_path: Path to the input Excel file (e.g., "data.xlsx")
    :param column_name: Name of the column to extract (e.g., "Name")
    :param output_path: Path to save the extracted data (e.g., "output.txt")
    :return: Boolean, True if success, False if fail
    """
    # Check if Excel file exists
    if not os.path.exists(excel_path):
        print(f"Error: Excel file {excel_path} does not exist!")
        return False
    
    try:
        # Load Excel workbook (read only mode for performance)
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        # Get the first worksheet
        ws = wb.active
        
        # Find the column index by column name
        column_index = None
        for cell in ws[1]:  # First row is header
            if cell.value == column_name:
                column_index = cell.column
                break
        
        if column_index is None:
            print(f"Error: Column {column_name} not found!")
            return False
        
        # Extract data from the column
        extracted_data = []
        # Skip header row (start from row 2)
        for row in ws.iter_rows(min_row=2, min_col=column_index, max_col=column_index):
            cell_value = row[0].value
            if cell_value is not None:  # Skip empty cells
                extracted_data.append(str(cell_value))
        
        # Save data to txt file
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(extracted_data))
        
        print(f"Success! Extracted {len(extracted_data)} rows to {output_path}")
        return True
    
    except Exception as e:
        print(f"Error during extraction: {str(e)}")
        return False

# Test the function
if __name__ == "__main__":
    # You can replace these paths with your own
    extract_excel_column("test_data.xlsx", "Name", "extracted_names.txt")
